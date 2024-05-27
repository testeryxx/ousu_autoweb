# -*- coding: utf-8 -*-

import openpyxl
from VAR.var import *
import yaml


class FileReader:
    """
    专门用来读取和写入yaml、excel文件
    """

    # 读取excel--openpyxl -- 文件格式：.xlsx
    @staticmethod  # 直接通过类名进行调用
    def read_excel(sheet_name, file_path=FILE_PATH):
        """
        读取Excel文件，只支持 .xlsx文件
        :param file_path: 文件路径
        :return: excel文件数据,元组的格式
        """
        # 打开现有的Excel文件或创建新的文件
        try:
            #  正常情况下直接打开
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # 选择或创建指定的工作表
        if sheet_name in workbook.sheetnames:
            # 【正常】  判断有没有对应的shtttname ，有的话把对应的数据给我加载出来
            worksheet = workbook[sheet_name]
        else:
            # 没有的话，给我新建一个
            worksheet = workbook.create_sheet(sheet_name)

        # 获取列名 --- 把第2行的数据拿出来作为我们的key值
        headers = [cell.value for cell in worksheet[1]]
        # print("所有的key", headers)

        # 将数据存储为字典,并且放在我们data当中
        data = []  # 所有的数据

        # 把小的数据从第三行开始
        for row in worksheet.iter_rows(min_row=3, values_only=True):
            # 把所有的数据直接加进去
            # data.append(dict(zip(headers, row)))

            #  当 is_true == True才应该加进去
            new_data = dict(zip(headers, row))
            if new_data["is_true"] is True:
                data.append(new_data)
        workbook.close()
        # 所有的数据
        return data

    @staticmethod
    def writeDataToExcel(sheet_name, file_path=FILE_PATH, row=None, column=None, value=None):
        # 打开现有的Excel文件或创建新的文件
        try:
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # 选择或创建指定的工作表
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(sheet_name)

        # 写入数据到指定行和列
        worksheet.cell(row=row, column=column).value = value

        # 保存修改后的文件--- 所以执行过程当中excel是要关闭的状态
        workbook.save(file_path)

    @staticmethod
    def read_yaml(file_path=YAML_PATH):
        """
        读取yaml文件
        :param file_path: 文件路径
        :return: yaml文件数据
        """
        with open(file_path, 'r', encoding="utf-8") as file:
            data = yaml.safe_load(file)
        return data

    @staticmethod
    def write_yaml(data, file_path=YAML_PATH):
        """
        写入yaml文件，写入无序没有关系，通过key获取数据
        :param data: 需要写入的数据
        :param file_path: 文件路径
        :return:
        """
        with open(file_path, 'w', encoding="utf-8") as file:
            # allow_unicode=True，避免将中文字符转换为 Unicode 转义序列
            yaml.dump(data, file, allow_unicode=True)

# if __name__ == '__main__':
#     res =FileReader.read_excel(sheet_name='baseinfo')
#     res[0]['项目名称']
#     print(res)
#     print(res[0]['项目名称'])
